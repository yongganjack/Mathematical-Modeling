"""问题三的变量边界以及可复用的、保护模板的 Excel 导出辅助函数。"""

from __future__ import annotations

import shutil
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from question1.data_processing import ProblemData, sha256_file
from question1.model import DerivedBomb, UAVPlan, derive_bomb, max_fuse_delay, missile_hit_time


TEMPLATE_HASHES = {
    "result1.xlsx": "af04b16e6a4719628971bcf5a03d230c9da6738e67eebac9276d254fdd4df1a7",
    "result2.xlsx": "c681d5e378538f71c77fca199a3ca8303a04dbcfc7bd95f870ae22f01ab69f91",
    "result3.xlsx": "b648c82d63e459ba6e6b3711ae79875e373521cd543b45571c4d8ff1ad5ec54a",
}

TEMPLATE_SPECS: dict[str, dict[str, Any]] = {
    "result1.xlsx": {
        "sheet": "Sheet1", "header_row": 1, "data_rows": (2, 3, 4),
        "fixed": {"C2": 1, "C3": 2, "C4": 3},
        "safe_columns": (1, 2, 4, 5, 6, 7, 8, 9, 10),
        "blank_row": 5, "note_cell": "A6",
    },
    "result2.xlsx": {
        "sheet": "Sheet1", "header_row": 1, "data_rows": (2, 3, 4),
        "fixed": {"A2": "FY1", "A3": "FY2", "A4": "FY3"},
        "safe_columns": tuple(range(2, 11)), "blank_row": 5, "note_cell": "B6",
    },
    "result3.xlsx": {
        "sheet": "Sheet1", "header_row": 1, "data_rows": tuple(range(2, 17)),
        "fixed": {
            **{f"A{row}": f"FY{(row - 2) // 3 + 1}" for row in range(2, 17)},
            **{f"D{row}": (row - 2) % 3 + 1 for row in range(2, 17)},
        },
        "safe_columns": (2, 3, 5, 6, 7, 8, 9, 10, 11, 12),
        "blank_row": 17, "note_cell": "B18",
    },
}


def q3_bounds(data: ProblemData) -> list[tuple[float, float]]:
    """返回 theta、速度、基准释放时刻、两个释放松弛量以及三个引信延迟的边界。"""

    hit = missile_hit_time(0, data)
    gap = float(data.min_release_interval)
    base_upper = max(0.0, hit - 2.0 * gap)
    fuse_upper = min(max_fuse_delay(0, data), hit)
    return [
        (-float(np.pi), float(np.pi)),
        tuple(map(float, data.uav_speed_bounds)),
        (0.0, base_upper),
        (0.0, base_upper),
        (0.0, base_upper),
        (0.0, fuse_upper),
        (0.0, fuse_upper),
        (0.0, fuse_upper),
    ]


def q3_config(config: Mapping[str, Any] | None) -> dict[str, int]:
    """解析可复现的计算预算，快速优化器约 80~140 次调用。"""

    cfg = config or {}
    runtime = cfg.get("optimization", {}).get("q3_runtime", {})
    if runtime:
        return {name: int(runtime[name]) for name in (
            "pso_particles", "pso_iterations", "de_particles", "de_iterations"
        )}
    requested = int(cfg.get("optimization", {}).get("budgets", {}).get("q3", {}).get("max_evaluations", 4000))
    if str(cfg.get("profile", "quick")).lower() == "quick":
        return {"pso_particles": 10, "pso_iterations": 8, "de_particles": 2, "de_iterations": 4}
    pso_particles = max(16, min(40, requested // 1000 + 16))
    return {
        "pso_particles": pso_particles,
        "pso_iterations": max(12, min(60, requested // pso_particles // 8)),
        "de_particles": max(4, min(12, requested // 8000 + 4)),
        "de_iterations": max(8, min(40, requested // 4000 + 8)),
    }


def inspect_excel_template(path: str | Path, template_name: str | None = None) -> dict[str, Any]:
    """读取用于保护官方结果模板的结构性值。"""

    source = Path(path)
    name = template_name or source.name
    if name not in TEMPLATE_SPECS:
        raise ValueError(f"不支持的 Excel 模板: {name}")
    spec = TEMPLATE_SPECS[name]
    workbook = load_workbook(source, data_only=False)
    if spec["sheet"] not in workbook.sheetnames:
        raise ValueError(f"模板缺少工作表 {spec['sheet']}")
    sheet = workbook[spec["sheet"]]
    max_column = max(spec["safe_columns"] + tuple(sheet[cell].column for cell in spec["fixed"]))
    header = [sheet.cell(spec["header_row"], col).value for col in range(1, max_column + 1)]
    return {
        "path": str(source.resolve()),
        "template_name": name,
        "sha256": sha256_file(source),
        "sheet_names": list(workbook.sheetnames),
        "header": header,
        "fixed": {cell: sheet[cell].value for cell in spec["fixed"]},
        "blank_row": [sheet.cell(spec["blank_row"], col).value for col in range(1, max_column + 1)],
        "note": sheet[spec["note_cell"]].value,
    }


def validate_excel_template(path: str | Path, template_name: str | None = None, *, require_official_hash: bool = False) -> dict[str, Any]:
    """验证表头、固定单元格、分隔行、注释以及可选的哈希值。"""

    info = inspect_excel_template(path, template_name)
    name = info["template_name"]
    spec = TEMPLATE_SPECS[name]
    errors: list[str] = []
    if any(value is None for value in info["header"]):
        errors.append("header contains blank cells")
    if info["fixed"] != spec["fixed"]:
        errors.append("fixed identifiers changed")
    if any(value is not None for value in info["blank_row"]):
        errors.append("separator row is not blank")
    if not isinstance(info["note"], str) or "0~360" not in info["note"]:
        errors.append("direction note changed")
    if require_official_hash and info["sha256"].lower() != TEMPLATE_HASHES[name]:
        errors.append("official template hash mismatch")
    return {**info, "valid": not errors, "errors": errors}


def copy_excel_template(source: str | Path, destination: str | Path, template_name: str | None = None) -> tuple[Path, str]:
    """复制经过验证的官方模板，且绝不覆盖源文件。"""

    source_path = Path(source).resolve()
    destination_path = Path(destination).resolve()
    if source_path == destination_path:
        raise ValueError("Excel 目标路径必须与官方模板不同")
    before = sha256_file(source_path)
    validation = validate_excel_template(source_path, template_name, require_official_hash=True)
    if not validation["valid"]:
        raise ValueError("官方 Excel 模板无效: " + "; ".join(validation["errors"]))
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    if destination_path.exists():
        raise FileExistsError(f"拒绝覆盖已有导出文件: {destination_path}")
    shutil.copy2(source_path, destination_path)
    if sha256_file(source_path) != before:
        raise RuntimeError("官方 Excel 模板在复制过程中被修改")
    return destination_path, before


def write_excel_rows(path: str | Path, template_name: str, rows: Sequence[Mapping[int | str, Any]], *, number_format: str = "0.000") -> None:
    """仅写入模板规范中声明为可安全修改的单元格。"""

    spec = TEMPLATE_SPECS[template_name]
    if len(rows) != len(spec["data_rows"]):
        raise ValueError("行数与模板不匹配")
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[spec["sheet"]]
    for row_number, values in zip(spec["data_rows"], rows):
        for column, value in values.items():
            column_number = column_index_from_string(column) if isinstance(column, str) else int(column)
            if column_number not in spec["safe_columns"]:
                raise ValueError(f"列 {column} 在 {template_name} 中受保护")
            cell = sheet.cell(row_number, column_number)
            if value is not None and (isinstance(value, (np.floating, np.integer)) or isinstance(value, (int, float)) and not isinstance(value, bool)):
                cell.value = float(value)
                cell.number_format = number_format
            else:
                cell.value = value
    workbook.save(path)


def read_excel_rows(path: str | Path, template_name: str) -> list[dict[int, Any]]:
    """将安全结果单元格读回为纯 Python 值。"""

    spec = TEMPLATE_SPECS[template_name]
    sheet = load_workbook(path, data_only=False)[spec["sheet"]]
    return [
        {column: sheet.cell(row, column).value for column in spec["safe_columns"]}
        for row in spec["data_rows"]
    ]


def export_result1(
    plan: UAVPlan,
    derived_bombs: Sequence[DerivedBomb | None] | None,
    data: ProblemData,
    template_path: str | Path,
    destination: str | Path,
    contributions: Sequence[float],
) -> tuple[Path, dict[str, Any]]:
    """将问题三的三弹方案导出到 result1.xlsx 的受保护副本中。"""

    if not isinstance(plan, UAVPlan) or len(plan.bombs) != 3:
        raise ValueError("result1 导出需要一个包含三枚炸弹的 UAV 方案")
    if len(contributions) != 3 or not np.all(np.isfinite(np.asarray(contributions, dtype=float))):
        raise ValueError("贡献值必须包含三个有限值")
    if derived_bombs is None or len(derived_bombs) != 3 or not all(isinstance(item, DerivedBomb) for item in derived_bombs):
        derived = [derive_bomb(plan, bomb, data) for bomb in plan.bombs]
    else:
        derived = list(derived_bombs)
    output, source_hash = copy_excel_template(template_path, destination, "result1.xlsx")
    heading_deg = float(np.degrees(plan.heading_rad) % 360.0)
    rows = []
    for bomb, contribution in zip(derived, contributions):
        rows.append({
            1: heading_deg, 2: float(plan.speed),
            4: float(bomb.release_point[0]), 5: float(bomb.release_point[1]), 6: float(bomb.release_point[2]),
            7: float(bomb.explosion_point[0]), 8: float(bomb.explosion_point[1]), 9: float(bomb.explosion_point[2]),
            10: float(contribution),
        })
    write_excel_rows(output, "result1.xlsx", rows)
    validation = validate_excel_template(output, "result1.xlsx", require_official_hash=False)
    readback = read_excel_rows(output, "result1.xlsx")
    numeric_ok = all(
        np.isclose(float(readback[index][column]), float(rows[index][column]), rtol=0.0, atol=1e-9)
        for index in range(3) for column in TEMPLATE_SPECS["result1.xlsx"]["safe_columns"]
    )
    source_after = sha256_file(template_path)
    if source_hash != source_after:
        raise RuntimeError("官方 result1.xlsx 在导出过程中被修改")
    validation.update({
        "source_hash_before": source_hash, "source_hash_after": source_after,
        "destination": str(output), "numeric_readback_valid": bool(numeric_ok),
        "readback": readback,
    })
    validation["valid"] = bool(validation["valid"] and numeric_ok and source_hash == source_after)
    if not validation["valid"]:
        raise RuntimeError("result1.xlsx 导出验证失败")
    return output, validation

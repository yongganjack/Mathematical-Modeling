from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from question1.data_processing import load_config, load_problem_data, sha256_file
from question1.model import BombPlan, UAVPlan
from question5.data_processing import export_result3
from question5.model import decode_route_particle, enumerate_lexicographic


PROJECT_DIR = Path(__file__).resolve().parents[1]


def test_integer_route_decoding_and_lexicographic_enumeration() -> None:
    decoded = decode_route_particle([-9.0, 0.49, 1.50, 99.0, 2.51], [2, 3, 4, 5, 4])
    assert decoded == (0, 0, 2, 4, 3)

    score_matrix = {
        (0, 0, 0): (10.0, 1.0), (0, 0, 1): (9.8, 4.0),
        (0, 1, 0): (9.8, 4.0), (0, 1, 1): (9.6, 5.0),
        (1, 0, 0): (9.9, 3.0), (1, 0, 1): (9.7, 5.0),
        (1, 1, 0): (9.7, 2.0), (1, 1, 1): (9.5, 6.0),
    }
    best_ids, best_score = enumerate_lexicographic(
        [2, 2, 2], lambda ids: score_matrix[ids], epsilon_J=0.2
    )
    assert best_ids == (0, 0, 1)
    assert best_score == (9.8, 4.0)


def test_result3_export_preserves_template_and_fixed_cells(tmp_path: Path) -> None:
    config = load_config(PROJECT_DIR / "configs" / "quick.json")
    data = load_problem_data(config)
    plans = [
        UAVPlan(0, 0.1, 100.0, (BombPlan(1, 1.0, 1.0, 0),)),
        UAVPlan(1, -0.2, 110.0, ()),
        UAVPlan(2, 0.3, 120.0, (BombPlan(1, 2.0, 1.0, 2),)),
        UAVPlan(3, -0.4, 90.0, ()),
        UAVPlan(4, 0.5, 80.0, ()),
    ]
    template = PROJECT_DIR / "00_赛题资料" / "附件" / "result3.xlsx"
    before = sha256_file(template)
    output, validation = export_result3(
        plans,
        data,
        template,
        tmp_path / "result3.xlsx",
        {(0, 1): 1.25, (2, 1): 0.75},
    )
    assert validation["valid"] is True
    assert sha256_file(template) == before

    sheet = load_workbook(output, data_only=False)["Sheet1"]
    assert [sheet[f"A{row}"].value for row in range(2, 17)] == [
        f"FY{(row - 2) // 3 + 1}" for row in range(2, 17)
    ]
    assert [sheet[f"D{row}"].value for row in range(2, 17)] == [
        (row - 2) % 3 + 1 for row in range(2, 17)
    ]
    assert all(sheet.cell(17, column).value is None for column in range(1, 13))
    assert isinstance(sheet["B18"].value, str) and "0~360" in sheet["B18"].value
    assert all(sheet.cell(3, column).value is None for column in (2, 3, 5, 6, 7, 8, 9, 10, 11, 12))

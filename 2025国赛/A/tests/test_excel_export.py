from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from question1.data_processing import load_config, load_problem_data, sha256_file
from question3.data_processing import export_result1
from question3.model import decode_q3_candidate
from question4.data_processing import export_result2
from question4.model import decode_q4_candidate


def test_export_result1_preserves_template_and_round_trips(tmp_path: Path) -> None:
    root = Path(__file__).resolve().parents[1]
    config = load_config(root / "configs" / "quick.json")
    data = load_problem_data(config)
    candidate = np.array([0.4, 100.0, 0.0, 0.0, 0.0, 1.0, 1.0, 1.0])
    plan = decode_q3_candidate(candidate, data)
    derived = [None]
    out, validation = export_result1(
        plan,
        derived,
        data,
        root / "00_赛题资料" / "附件" / "result1.xlsx",
        tmp_path / "result1.xlsx",
        contributions=[1.25, 2.5, 3.75],
    )
    assert Path(out).resolve() == (tmp_path / "result1.xlsx").resolve()
    assert validation["valid"] is True
    ws = load_workbook(out, data_only=False)["Sheet1"]
    assert [ws.cell(r, 3).value for r in (2, 3, 4)] == [1, 2, 3]
    assert ws["A6"].value and "x轴" in ws["A6"].value
    assert ws["A5"].value is None
    assert isinstance(ws["A2"].value, float)


def test_decode_q3_enforces_release_spacing(problem_data) -> None:
    plan = decode_q3_candidate([0.0, 90.0, 2.0, -10.0, 0.1, 1.0, 2.0, 0.0], problem_data)
    releases = [bomb.release_time for bomb in plan.bombs]
    assert all(second - first >= 1.0 - 1e-9 for first, second in zip(releases, releases[1:]))


def test_decode_q4_returns_three_single_bomb_plans(problem_data) -> None:
    plans = decode_q4_candidate(np.arange(12, dtype=float), problem_data)
    assert [plan.uav_index for plan in plans] == [0, 1, 2]
    assert [(plan.bombs[0].bomb_index, plan.bombs[0].assigned_missile) for plan in plans] == [(1, 0)] * 3


def test_export_result2_preserves_fixed_template_cells(tmp_path: Path) -> None:
    root = Path(__file__).resolve().parents[1]
    data = load_problem_data(load_config(root / "configs" / "quick.json"))
    plans = decode_q4_candidate([0.0, 100.0, 1.0, 2.0] * 3, data)
    source = root / "00_赛题资料" / "附件" / "result2.xlsx"
    before = sha256_file(source)
    out, validation = export_result2(plans, None, data, source, tmp_path / "result2.xlsx", [1.0, 2.0, 3.0])
    ws = load_workbook(out, data_only=False)["Sheet1"]
    assert validation["valid"] and sha256_file(source) == before
    assert [ws[f"A{row}"].value for row in (2, 3, 4)] == ["FY1", "FY2", "FY3"]
    assert all(ws.cell(5, column).value is None for column in range(1, 11))
    assert ws["B6"].value and "0~360" in ws["B6"].value
